"""Shared data loaders and derivation helpers for the Stay-In School pipeline."""
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import numpy as np
import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
ARTIFACTS = ROOT / "artifacts"
ARTIFACTS.mkdir(exist_ok=True)

CSV_2023 = DATA / "data_FIN_YEAR_2023-2024.csv"
CSV_2024 = DATA / "data_FIN_YEAR_2024-2025.csv"
DROPPED_2023 = DATA / "CHILDSNO_Dropped_2023_24.xlsx"
DROPPED_2024 = DATA / "CHILDSNO_Dropped_2024_25.xlsx"

META_COLS = ["schoolid", "GENDER", "CASTE", "DOB", "CHILD_SNO", "FIN_YEAR"]
MARK_COLS = ["FA1_MARKS", "FA2_MARKS", "FA3_MARKS", "FA4_MARKS", "SA1_MARKS", "SA2_MARKS"]

# Month-based grouping of the academic calendar (AP school year: Jun -> Apr)
MONTH_ORDER = ["Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr"]

DAY_RE = re.compile(r"^\d{2}-[A-Za-z]{3}$")


def _clean_nan(o):
    """Recursively convert NaN / inf / numpy scalars to JSON-valid values."""
    if isinstance(o, float):
        return None if (np.isnan(o) or np.isinf(o)) else o
    if isinstance(o, (np.floating,)):
        v = float(o)
        return None if (np.isnan(v) or np.isinf(v)) else v
    if isinstance(o, (np.integer,)):
        return int(o)
    if isinstance(o, np.ndarray):
        return _clean_nan(o.tolist())
    if isinstance(o, (pd.Timestamp, datetime)):
        return o.isoformat()
    if isinstance(o, dict):
        return {k: _clean_nan(v) for k, v in o.items()}
    if isinstance(o, (list, tuple)):
        return [_clean_nan(x) for x in o]
    return o


def write_json(path: Path, obj) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    cleaned = _clean_nan(obj)
    with path.open("w") as f:
        # allow_nan=False forces us to emit valid JSON; _clean_nan has already removed NaN/inf
        json.dump(cleaned, f, separators=(",", ":"), allow_nan=False, default=_default)


def _default(o):
    if isinstance(o, (np.integer,)):
        return int(o)
    if isinstance(o, (np.floating,)):
        v = float(o)
        return None if (np.isnan(v) or np.isinf(v)) else v
    if isinstance(o, np.ndarray):
        return _clean_nan(o.tolist())
    if isinstance(o, (pd.Timestamp, datetime)):
        return o.isoformat()
    raise TypeError(f"not serializable: {type(o)}")


def day_columns(df: pd.DataFrame) -> list[str]:
    return [c for c in df.columns if DAY_RE.match(c)]


def load_dropped(path: Path) -> set[int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: set[int] = set()
    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            v = row[0]
            if v is None:
                continue
            try:
                out.add(int(v))
            except (TypeError, ValueError):
                continue
    return out


def _parse_dob(s: str) -> pd.Timestamp | pd._libs.tslibs.nattype.NaTType:
    if not isinstance(s, str) or not s.strip():
        return pd.NaT
    # accept dd/mm/yy or dd/mm/yyyy, plus ISO
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            ts = pd.to_datetime(s, format=fmt, errors="raise")
            # dd/mm/yy may interpret 08 as 2008 (correct) or 68 as 2068; normalize to <=2016
            if fmt == "%d/%m/%y" and ts.year > 2016:
                ts = ts.replace(year=ts.year - 100)
            return ts
        except Exception:
            continue
    return pd.NaT


def derive_geo(schoolid_series: pd.Series) -> pd.DataFrame:
    """Extract district / block / school_seq from the AP DISE-style schoolid.

    AP school codes are typically 11 digits: SS DD BBB SSSS (state, district, block, school).
    We zero-pad to 11 and slice. The STATE digits are used only for sanity.
    """
    s = schoolid_series.astype("string").fillna("").str.replace(".0", "", regex=False)
    s = s.str.zfill(11)
    return pd.DataFrame({
        "schoolid": schoolid_series.astype("string"),
        "state_code": s.str[0:2],
        "district_code": s.str[2:4],
        "block_code": s.str[4:7],
        "school_seq": s.str[7:11],
    })


# ---------- AP district codes ----------
# Post-2014 Andhra Pradesh: state code 28 historically refers to combined AP/TG DISE; the
# 13-district AP map below covers the districts present in the uploaded datasets. Codes
# that don't match a known district fall back to "District {code}" which keeps the
# dashboard honest without fabricating identities.
AP_DISTRICT_NAMES: dict[str, str] = {
    "10": "Srikakulam",
    "11": "Vizianagaram",
    "12": "Visakhapatnam",
    "13": "East Godavari",
    "14": "West Godavari",
    "15": "Krishna",
    "16": "Guntur",
    "17": "Prakasam",
    "18": "SPS Nellore",
    "19": "YSR Kadapa",
    "20": "Kurnool",
    "21": "Anantapuramu",
    "22": "Chittoor",
}


def district_name(code: str) -> str:
    return AP_DISTRICT_NAMES.get(code, f"District {code}")


# ---------- main loader ----------
@dataclass
class LoadedYear:
    year: str
    meta: pd.DataFrame       # student_id index + demographic + geo
    att: np.ndarray          # shape (N, D) values in {1, 0, NaN} with np.nan for unknown
    day_cols: list[str]
    marks: pd.DataFrame      # numeric marks
    dropped: np.ndarray      # bool per student


def load_year(csv_path: Path, dropped_path: Path, year: str,
              nrows: int | None = None) -> LoadedYear:
    df = pd.read_csv(csv_path, dtype=str, nrows=nrows, low_memory=False)
    df.columns = [c.strip() for c in df.columns]
    days = day_columns(df)

    # --- attendance as float matrix ---
    att = df[days].to_numpy(copy=False)
    # map values to 1/0/nan
    mat = np.full(att.shape, np.nan, dtype=np.float32)
    mat[att == "Y"] = 1.0
    mat[att == "N"] = 0.0

    # --- marks ---
    marks = df[MARK_COLS].apply(pd.to_numeric, errors="coerce")

    # --- meta + geo ---
    geo = derive_geo(df["schoolid"])
    dob = df["DOB"].apply(_parse_dob)
    # reference date: end of academic year (30 Apr)
    ref = pd.Timestamp(int(year.split("-")[1]), 4, 30) if "-" in year else pd.Timestamp(2024, 4, 30)
    age = (ref - dob).dt.days / 365.25

    meta = pd.DataFrame({
        "child_sno": pd.to_numeric(df["CHILD_SNO"], errors="coerce").astype("Int64"),
        "schoolid": df["schoolid"].astype("string"),
        "gender": pd.to_numeric(df["GENDER"], errors="coerce").astype("Int64"),
        "caste": pd.to_numeric(df["CASTE"], errors="coerce").astype("Int64"),
        "dob": dob,
        "age_years": age.astype("float32"),
        "fin_year": year,
        "district_code": geo["district_code"],
        "block_code": geo["block_code"],
        "school_seq": geo["school_seq"],
    })

    # --- dropped flag ---
    dropped_ids = load_dropped(dropped_path)
    dropped = meta["child_sno"].isin(dropped_ids).fillna(False).to_numpy()

    return LoadedYear(year=year, meta=meta, att=mat, day_cols=days,
                      marks=marks.astype("float32"), dropped=dropped)


# ---------- feature computation ----------
def month_of(col: str) -> str:
    return col.split("-")[1]


def month_indices(day_cols: Iterable[str]) -> dict[str, list[int]]:
    idx: dict[str, list[int]] = {m: [] for m in MONTH_ORDER}
    for i, c in enumerate(day_cols):
        m = month_of(c)
        if m in idx:
            idx[m].append(i)
    return idx


# ---------- stat helpers ----------
def rolling_min_rate(att: np.ndarray, window: int) -> np.ndarray:
    """Min attendance rate over any rolling window of `window` tracked days per student.

    Tracks only non-NaN days. Uses a cumulative-sum sliding approach and returns the
    minimum rate achieved across the year — a proxy for the student's worst run.
    """
    mask = ~np.isnan(att)
    pres = np.where(mask, att, 0.0)
    # valid counts per window
    csum_pres = np.concatenate([np.zeros((att.shape[0], 1), dtype=np.float32),
                                np.cumsum(pres, axis=1, dtype=np.float32)], axis=1)
    csum_mask = np.concatenate([np.zeros((att.shape[0], 1), dtype=np.float32),
                                np.cumsum(mask.astype(np.float32), axis=1)], axis=1)
    w = window
    total_pres = csum_pres[:, w:] - csum_pres[:, :-w]
    total_mask = csum_mask[:, w:] - csum_mask[:, :-w]
    with np.errstate(invalid="ignore", divide="ignore"):
        rates = np.where(total_mask > 0, total_pres / np.maximum(total_mask, 1), np.nan)
    # ignore windows with too few tracked days
    rates = np.where(total_mask >= max(3, w // 2), rates, np.nan)
    return np.nanmin(rates, axis=1)


def longest_absence_streak(att: np.ndarray) -> np.ndarray:
    """Longest contiguous run of N values (treating NaN as neutral)."""
    absent = (att == 0.0)
    out = np.zeros(att.shape[0], dtype=np.int32)
    run = np.zeros(att.shape[0], dtype=np.int32)
    for j in range(att.shape[1]):
        col = absent[:, j]
        run = np.where(col, run + 1, 0)
        out = np.maximum(out, run)
    return out


def absence_streak_count(att: np.ndarray, min_len: int = 3) -> np.ndarray:
    """Count number of absence streaks >= min_len days."""
    absent = (att == 0.0)
    n, d = absent.shape
    out = np.zeros(n, dtype=np.int32)
    run = np.zeros(n, dtype=np.int32)
    prev = np.zeros(n, dtype=bool)
    for j in range(d):
        col = absent[:, j]
        run = np.where(col, run + 1, 0)
        ended = (~col) & prev & (np.roll(run, 0) == 0)  # simple terminator
        # increment count where a streak just ended and it was >= min_len
        # we detect by comparing previous run length to threshold when col flips to False
        prev = col
    # re-implement more simply:
    out[:] = 0
    run = np.zeros(n, dtype=np.int32)
    for j in range(d):
        col = absent[:, j]
        run = np.where(col, run + 1, 0)
        if j > 0:
            just_ended = (~col) & (absent[:, j - 1])
            out += ((run == 0) & just_ended & (_prev_run_len(absent, j - 1) >= min_len)).astype(np.int32)
    # final: if streak ran to the end
    tail = (run >= min_len)
    out += tail.astype(np.int32)
    return out


def _prev_run_len(absent: np.ndarray, idx: int) -> np.ndarray:
    """Helper: length of trailing absence run ending at column idx (inclusive)."""
    n = absent.shape[0]
    out = np.zeros(n, dtype=np.int32)
    k = idx
    still = absent[:, k].copy()
    while k >= 0 and still.any():
        out += still.astype(np.int32)
        k -= 1
        if k < 0:
            break
        still = still & absent[:, k]
    return out


def recent_deterioration(att: np.ndarray, window: int = 30) -> np.ndarray:
    """Delta between early-year and late-year attendance rates."""
    mask = ~np.isnan(att)
    d = att.shape[1]
    w = min(window, d // 2)
    early = _rate(att[:, :w], mask[:, :w])
    late = _rate(att[:, -w:], mask[:, -w:])
    return early - late  # positive = deterioration


def _rate(att_slice: np.ndarray, mask_slice: np.ndarray) -> np.ndarray:
    valid = mask_slice.sum(axis=1)
    pres = np.nansum(att_slice, axis=1)
    with np.errstate(invalid="ignore", divide="ignore"):
        rate = np.where(valid > 0, pres / np.maximum(valid, 1), np.nan)
    return rate.astype(np.float32)


def window_rate(att: np.ndarray, start: int, end: int) -> np.ndarray:
    mask = ~np.isnan(att[:, start:end])
    return _rate(att[:, start:end], mask)
